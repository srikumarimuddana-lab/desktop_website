#!/usr/bin/env python3
"""
Generate an Excel file from the knowledge base seed SQL files
for human review and validation.
"""

import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SEEDS_DIR = "/home/user/desktop_website/supabase/seeds"
OUTPUT = "/home/user/desktop_website/supabase/spinr_knowledge_base_review.xlsx"

FILES = [
    ("01_general_info.sql", "General Info"),
    ("02_rider_getting_started.sql", "Rider - Getting Started"),
    ("03_rider_payments_cancellation.sql", "Rider - Payments & Cancellation"),
    ("04_rider_safety_accessibility.sql", "Rider - Safety & Accessibility"),
    ("05_driver_getting_started.sql", "Driver - Getting Started"),
    ("06_driver_earnings_subscriptions.sql", "Driver - Earnings & Subscriptions"),
    ("07_driver_tax_operations.sql", "Driver - Tax & Operations"),
    ("08_account_management.sql", "Account Management"),
    ("09_app_troubleshooting.sql", "App & Troubleshooting"),
    ("10_policies_legal_seasonal.sql", "Policies, Legal & Seasonal"),
]

def parse_sql_entries(filepath):
    """Parse INSERT INTO statements and extract title, content, category, tags."""
    with open(filepath, 'r') as f:
        sql = f.read()

    entries = []
    # Match each tuple block: ( 'title', 'content', 'category', ARRAY[...], 'source' )
    pattern = re.compile(
        r"\(\s*\n\s*'((?:[^']|'')*?)'\s*,\s*\n\s*'((?:[^']|'')*?)'\s*,\s*\n\s*'((?:[^']|'')*?)'\s*,\s*\n\s*ARRAY\[(.*?)\]\s*,\s*\n\s*'website_analysis'\s*\n\s*\)",
        re.DOTALL
    )

    for m in pattern.finditer(sql):
        title = m.group(1).replace("''", "'")
        content = m.group(2).replace("''", "'")
        category = m.group(3).replace("''", "'")
        tags_raw = m.group(4)
        tags = ", ".join(t.strip().strip("'") for t in tags_raw.split(","))
        entries.append((title, content, category, tags))

    return entries


def create_excel():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Colors
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    even_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    approved_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    rejected_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)

    # ========================
    # SUMMARY SHEET
    # ========================
    ws_summary = wb.create_sheet("Summary")
    ws_summary.sheet_properties.tabColor = "1F4E79"

    summary_headers = ["#", "Category", "Sheet Name", "Entry Count", "Approved", "Needs Edit", "Rejected"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    total = 0
    all_entries = []

    for idx, (filename, sheet_name) in enumerate(FILES):
        filepath = os.path.join(SEEDS_DIR, filename)
        entries = parse_sql_entries(filepath)
        count = len(entries)
        total += count
        all_entries.extend([(sheet_name, e) for e in entries])

        row = idx + 2
        ws_summary.cell(row=row, column=1, value=idx + 1).alignment = center_align
        ws_summary.cell(row=row, column=2, value=sheet_name).alignment = wrap_align
        ws_summary.cell(row=row, column=3, value=filename).alignment = wrap_align
        ws_summary.cell(row=row, column=4, value=count).alignment = center_align
        ws_summary.cell(row=row, column=5, value="").alignment = center_align  # Approved
        ws_summary.cell(row=row, column=6, value="").alignment = center_align  # Needs Edit
        ws_summary.cell(row=row, column=7, value="").alignment = center_align  # Rejected
        for c in range(1, 8):
            ws_summary.cell(row=row, column=c).border = thin_border
            if idx % 2 == 1:
                ws_summary.cell(row=row, column=c).fill = even_fill

    # Total row
    total_row = len(FILES) + 2
    ws_summary.cell(row=total_row, column=1, value="").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True, size=12)
    ws_summary.cell(row=total_row, column=4, value=total).font = Font(bold=True, size=12)
    ws_summary.cell(row=total_row, column=4).alignment = center_align
    for c in range(1, 8):
        ws_summary.cell(row=total_row, column=c).border = thin_border

    ws_summary.column_dimensions['A'].width = 5
    ws_summary.column_dimensions['B'].width = 35
    ws_summary.column_dimensions['C'].width = 40
    ws_summary.column_dimensions['D'].width = 14
    ws_summary.column_dimensions['E'].width = 12
    ws_summary.column_dimensions['F'].width = 12
    ws_summary.column_dimensions['G'].width = 12

    # ========================
    # ALL ENTRIES SHEET (flat view)
    # ========================
    ws_all = wb.create_sheet("All Entries")
    ws_all.sheet_properties.tabColor = "2E75B6"

    all_headers = ["#", "Category", "Title (Question)", "Content (Answer)", "DB Category", "Tags", "Status", "Reviewer Notes"]
    for col, h in enumerate(all_headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, (sheet_name, (title, content, category, tags)) in enumerate(all_entries):
        row = i + 2
        ws_all.cell(row=row, column=1, value=i + 1).alignment = center_align
        ws_all.cell(row=row, column=2, value=sheet_name).alignment = wrap_align
        ws_all.cell(row=row, column=3, value=title).alignment = wrap_align
        ws_all.cell(row=row, column=4, value=content).alignment = wrap_align
        ws_all.cell(row=row, column=5, value=category).alignment = center_align
        ws_all.cell(row=row, column=6, value=tags).alignment = wrap_align
        ws_all.cell(row=row, column=7, value="").alignment = center_align  # Status dropdown placeholder
        ws_all.cell(row=row, column=8, value="").alignment = wrap_align  # Notes

        for c in range(1, 9):
            ws_all.cell(row=row, column=c).border = thin_border
            if i % 2 == 1:
                ws_all.cell(row=row, column=c).fill = even_fill

    # Add data validation for Status column
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Approved,Needs Edit,Rejected,Pending"', allow_blank=True)
    dv.error = "Please select: Approved, Needs Edit, Rejected, or Pending"
    dv.errorTitle = "Invalid Status"
    ws_all.add_data_validation(dv)
    dv.add(f"G2:G{len(all_entries) + 1}")

    ws_all.column_dimensions['A'].width = 5
    ws_all.column_dimensions['B'].width = 30
    ws_all.column_dimensions['C'].width = 45
    ws_all.column_dimensions['D'].width = 80
    ws_all.column_dimensions['E'].width = 12
    ws_all.column_dimensions['F'].width = 40
    ws_all.column_dimensions['G'].width = 14
    ws_all.column_dimensions['H'].width = 40

    # Freeze top row
    ws_all.freeze_panes = "A2"
    # Auto-filter
    ws_all.auto_filter.ref = f"A1:H{len(all_entries) + 1}"

    # ========================
    # CATEGORY-SPECIFIC SHEETS
    # ========================
    category_colors = [
        "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
        "70AD47", "264478", "9B59B6", "E74C3C", "1ABC9C"
    ]

    for idx, (filename, sheet_name) in enumerate(FILES):
        filepath = os.path.join(SEEDS_DIR, filename)
        entries = parse_sql_entries(filepath)

        # Truncate sheet name to 31 chars (Excel limit)
        safe_name = sheet_name[:31]
        ws = wb.create_sheet(safe_name)
        ws.sheet_properties.tabColor = category_colors[idx % len(category_colors)]

        cat_headers = ["#", "Title (Question)", "Content (Answer)", "Tags", "Status", "Reviewer Notes"]
        for col, h in enumerate(cat_headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for i, (title, content, category, tags) in enumerate(entries):
            row = i + 2
            ws.cell(row=row, column=1, value=i + 1).alignment = center_align
            ws.cell(row=row, column=2, value=title).alignment = wrap_align
            ws.cell(row=row, column=3, value=content).alignment = wrap_align
            ws.cell(row=row, column=4, value=tags).alignment = wrap_align
            ws.cell(row=row, column=5, value="").alignment = center_align
            ws.cell(row=row, column=6, value="").alignment = wrap_align

            for c in range(1, 7):
                ws.cell(row=row, column=c).border = thin_border
                if i % 2 == 1:
                    ws.cell(row=row, column=c).fill = even_fill

        # Data validation for status
        dv2 = DataValidation(type="list", formula1='"Approved,Needs Edit,Rejected,Pending"', allow_blank=True)
        ws.add_data_validation(dv2)
        dv2.add(f"E2:E{len(entries) + 1}")

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 40

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:F{len(entries) + 1}"

    # Save
    wb.save(OUTPUT)
    print(f"Excel file created: {OUTPUT}")
    print(f"Total entries: {total}")
    print(f"Sheets: Summary + All Entries + {len(FILES)} category sheets")


if __name__ == "__main__":
    create_excel()